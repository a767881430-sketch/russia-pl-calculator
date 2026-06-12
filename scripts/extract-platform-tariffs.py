from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DESKTOP = Path.home() / "Desktop"
OUTPUT = ROOT / "src" / "lib" / "platformTariffs.js"
JSON_OUTPUT = ROOT / "public" / "platform-tariffs.json"


def pick_file(predicate):
    matches = [path for path in DESKTOP.glob("*.xlsx*") if predicate(path)]
    if not matches:
        raise FileNotFoundError("No matching Excel file found on Desktop")
    return max(matches, key=lambda path: path.stat().st_mtime)


def pick_workbook(predicate, min_sheets):
    matches = []
    for path in DESKTOP.glob("*.xlsx*"):
        if not predicate(path):
            continue
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            if len(wb.worksheets) >= min_sheets:
                matches.append(path)
        except Exception:
            continue
    if not matches:
        raise FileNotFoundError("No matching workbook with required sheets found on Desktop")
    return max(matches, key=lambda path: path.stat().st_mtime)


def num(value, fallback=0):
    try:
        if value is None or value == "":
            return fallback
        if isinstance(value, str):
            value = value.replace(",", ".")
        return float(value)
    except (TypeError, ValueError):
        return fallback


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return " ".join(value.split())
    return value


def write_platform_tariffs(data):
    JSON_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    JSON_OUTPUT.write_text(
        json.dumps(data, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(
        "const EMPTY_PLATFORM_TARIFFS = {\n"
        "  generatedAt: \"\",\n"
        "  sources: {},\n"
        "  ozon: {\n"
        "    priceBuckets: [0, 100.01, 300.01, 1500.01, 5000.01, 10000.01],\n"
        "    rfbsPriceBuckets: [0, 1500.01, 5000.01, 10000.01],\n"
        "    lastMileRate: 0,\n"
        "    lastMileMin: 25,\n"
        "    lastMileMax: 25,\n"
        "    commissions: [],\n"
        "    freight: { volumeBands: [], clusters: [], rows: [] },\n"
        "    nonLocal: [],\n"
        "    clusters: [],\n"
        "  },\n"
        "  wb: {\n"
        "    commissions: [],\n"
        "    localization: [],\n"
        "    volumeRates: [],\n"
        "  },\n"
        "  yandex: {\n"
        "    paymentFrequencies: [\n"
        "      [\"每月一次\", 0.013],\n"
        "      [\"每两周一次\", 0.019],\n"
        "      [\"每周一次\", 0.022],\n"
        "      [\"每天\", 0.027],\n"
        "    ],\n"
        "    commissions: [],\n"
        "  },\n"
        "};\n\n"
        "const LOADED_PLATFORM_TARIFFS = globalThis.__PLATFORM_TARIFFS__ || globalThis.window?.__PLATFORM_TARIFFS__;\n\n"
        "export const PLATFORM_TARIFFS = LOADED_PLATFORM_TARIFFS || EMPTY_PLATFORM_TARIFFS;\n"
        "export const PLATFORM_TARIFFS_LOADED = !!LOADED_PLATFORM_TARIFFS;\n",
        encoding="utf-8",
    )


def extract_ozon(path: Path):
    wb = load_workbook(path, data_only=True, read_only=True)
    freight_ws = wb.worksheets[2]
    non_local_ws = wb.worksheets[3]
    commission_ws = wb.worksheets[4]

    commissions = []
    for row in commission_ws.iter_rows(min_row=3, values_only=True):
        product_type = clean(row[1])
        if not product_type:
            continue
        commissions.append([
            product_type,
            [num(v) for v in row[2:8]],
            [num(v) for v in row[14:20]],
            [num(v) for v in row[20:24]],
        ])

    freight = []
    volume_band_map = {}
    clusters = []
    cluster_index = {}

    def cluster_id(value):
        if value not in cluster_index:
            cluster_index[value] = len(clusters)
            clusters.append(value)
        return cluster_index[value]

    for row in freight_ws.iter_rows(min_row=2, values_only=True):
        if not row[1] or not row[2] or not row[3]:
            continue
        volume_text = clean(row[1])
        supply = clean(row[2])
        delivery = clean(row[3])
        if volume_text not in volume_band_map:
            volume_band_map[volume_text] = [num(row[0]), volume_text]
        freight.append([volume_text, cluster_id(supply), cluster_id(delivery), num(row[5])])

    volume_bands = sorted(volume_band_map.values(), key=lambda item: item[0])
    volume_index = {item[1]: index for index, item in enumerate(volume_bands)}
    freight = [[volume_index[row[0]], row[1], row[2], row[3]] for row in freight]

    non_local = []
    for row in non_local_ws.iter_rows(min_row=2, values_only=True):
        cluster = clean(row[0])
        if cluster:
            non_local.append([cluster, num(row[1])])

    return {
        "priceBuckets": [0, 100.01, 300.01, 1500.01, 5000.01, 10000.01],
        "rfbsPriceBuckets": [0, 1500.01, 5000.01, 10000.01],
        "lastMileRate": 0,
        "lastMileMin": 25,
        "lastMileMax": 25,
        "commissions": commissions,
        "freight": {
            "volumeBands": volume_bands,
            "clusters": clusters,
            "rows": freight,
        },
        "nonLocal": non_local,
        "clusters": sorted(clusters),
    }


def extract_wb(path: Path):
    wb = load_workbook(path, data_only=True, read_only=True)
    commission_ws = wb.worksheets[1]
    localization_ws = wb.worksheets[2]

    commissions = []
    for row in commission_ws.iter_rows(min_row=2, values_only=True):
        subcategory = clean(row[1])
        if not subcategory:
            continue
        commissions.append([
            subcategory,
            num(row[2]),
            num(row[3]),
            num(row[4]),
            num(row[5]),
        ])

    localization = []
    volume_rates = []
    for row in localization_ws.iter_rows(min_row=2, values_only=True):
        band = clean(row[0])
        if band:
            localization.append([band, num(row[1], 1), num(row[2])])
        volume_band = clean(row[5])
        if volume_band:
            volume_rates.append([volume_band, num(row[6])])

    return {
        "commissions": commissions,
        "localization": localization,
        "volumeRates": volume_rates,
    }


def extract_yandex_commissions(path: Path):
    wb = load_workbook(path, data_only=True, read_only=True)
    by_path = {}
    sheet_map = [
        ("fby", 0),
        ("fbs", 1),
        ("express", 2),
        ("dbs", 3),
    ]
    for key, index in sheet_map:
        ws = wb.worksheets[index]
        for row in ws.iter_rows(min_row=2, values_only=True):
            parts = [clean(v) for v in row[:7] if clean(v)]
            if not parts:
                continue
            path_name = " / ".join(parts)
            rates = by_path.setdefault(path_name, {"fby": None, "fbs": None, "express": None, "dbs": None})
            rates[key] = num(row[7])

    categories = []
    for path_name, rates in sorted(by_path.items(), key=lambda item: item[0]):
        categories.append([
            path_name,
            rates["fby"],
            rates["fbs"],
            rates["express"],
            rates["dbs"],
        ])

    return {
        "paymentFrequencies": [
            ["每月一次", 0.013],
            ["每两周一次", 0.019],
            ["每周一次", 0.022],
            ["每天", 0.027],
        ],
        "commissions": categories,
    }


def main():
    ozon_file = pick_workbook(lambda p: "OZON" in p.name and p.stat().st_size > 1_000_000, 5)
    wb_file = pick_workbook(lambda p: "WB" in p.name and p.stat().st_size > 1_000_000, 3)
    yandex_commission_file = pick_file(lambda p: p.name.startswith("Yandex-4"))

    data = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sources": {
            "ozon": ozon_file.name,
            "wb": wb_file.name,
            "yandex": yandex_commission_file.name,
        },
        "ozon": extract_ozon(ozon_file),
        "wb": extract_wb(wb_file),
        "yandex": extract_yandex_commissions(yandex_commission_file),
    }

    write_platform_tariffs(data)
    print(f"Wrote {JSON_OUTPUT} from {ozon_file.name}, {wb_file.name}, {yandex_commission_file.name}")


if __name__ == "__main__":
    main()
