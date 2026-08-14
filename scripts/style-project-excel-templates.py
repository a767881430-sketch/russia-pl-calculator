from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
FILES = [
    ROOT / "public" / "投资损益计算器完整项目填写模板.xlsx",
    ROOT / "public" / "A009-假发完整项目导入.xlsx",
    ROOT / "public" / "project-import-template.xlsx",
    ROOT / "public" / "a009-wig-import.xlsx",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
INPUT_FILL = PatternFill("solid", fgColor="EAF3F8")
STRUCTURE_FILL = PatternFill("solid", fgColor="F2F2F2")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10, color="1F2937")
INPUT_FONT = Font(name="Arial", size=10, color="0F4C81")
THIN_GRAY = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
PARAMETER_DROPDOWNS = {
    "税制": [
        "USN 6%（销售额税）",
        "USN 15%（利润税）",
        "USN 6% + 增值税 5%",
        "USN 6% + 增值税 7%",
        "USN 15% + 增值税 5%",
        "USN 15% + 增值税 7%",
        "АУСН 8%（销售额税）",
        "一般税制",
        "自定义税率",
    ],
    "收入税计税基数": ["按销售额计税", "按回款计税"],
    "计税收入确认方式": ["按含税销售额确认", "按平台结算单确认"],
    "可扣成本口径": ["按已取得凭证的账面成本", "按实际到仓成本（内部测算）"],
    "平台综合费可扣除": ["是", "否"],
    "回款损耗可扣除": ["是", "否"],
    "回款损耗计提基数": ["按标价销售额计提", "按实际回款计提"],
    "启用 ИП 附加保险": ["是", "否"],
    "自动增值税升级": ["是", "否"],
}


def apply_parameter_dropdowns(sheet) -> None:
    rows_by_name = {
        str(sheet.cell(row, 1).value or "").strip(): row
        for row in range(2, sheet.max_row + 1)
    }
    sheet.data_validations.dataValidation = []
    for name, options in PARAMETER_DROPDOWNS.items():
        row = rows_by_name.get(name)
        if not row:
            raise ValueError(f"项目参数工作表缺少下拉参数行：{name}")
        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(options)}"',
            allow_blank=False,
        )
        validation.errorStyle = "stop"
        validation.errorTitle = "请从下拉选项中选择"
        validation.error = "此项请使用下拉选项，不要自行填写。"
        validation.showErrorMessage = True
        validation.showInputMessage = True
        validation.promptTitle = name
        validation.prompt = "请从下拉选项中选择。"
        sheet.add_data_validation(validation)
        validation.add(sheet.cell(row, 2))


def apply_style(path: Path) -> None:
    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = True
        sheet.row_dimensions[1].height = 24

        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.font = INPUT_FONT
                cell.fill = INPUT_FILL

        if sheet.title == "项目参数":
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, 1).fill = STRUCTURE_FILL
                sheet.cell(row, 1).font = BODY_FONT
                sheet.cell(row, 3).fill = STRUCTURE_FILL
                sheet.cell(row, 3).font = BODY_FONT
                if sheet.cell(row, 1).value == "预测起始月份（YYYY-MM）":
                    sheet.cell(row, 2).number_format = "@"
            apply_parameter_dropdowns(sheet)

        for column in range(1, sheet.max_column + 1):
            max_length = max(
                (len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1)),
                default=10,
            )
            sheet.column_dimensions[get_column_letter(column)].width = min(max(max_length + 2, 12), 36)

    workbook.save(path)
    print(f"已美化 {path.name}")


for workbook_path in FILES:
    apply_style(workbook_path)
