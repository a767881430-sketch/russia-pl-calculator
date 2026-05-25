from __future__ import annotations

from html import escape
from pathlib import Path
import json
import re

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PUBLIC = ROOT / "public" / "deli-proposal"
ASSET_DIR = PUBLIC / "assets" / "ozon_product_images"
SOURCE_XLSX = (
    Path("D:/")
    / "妙成"
    / "星哈酷"
    / "企划书"
    / "重启得力玻璃"
    / "重庆得力玻璃类目趋势.xlsx"
)
OUT_HTML = PUBLIC / "ozon-competitors.html"

RUB_PER_RMB = 12.8


CATEGORY_LABELS = [
    ("Стакан", "水杯 / 玻璃杯", "透明玻璃款"),
    ("Бокал", "高脚杯 / 酒杯", "透明玻璃款"),
    ("Крышка для кружки, чашки, стакана", "杯盖 / 配件", "配件误差"),
    ("Посуда для чайных церемоний", "茶道器具", "参考观察"),
    ("Чайник заварочный", "玻璃茶壶", "透明玻璃款"),
    ("Кувшин", "水壶 / 冷水壶", "重点观察"),
    ("Сахарница", "糖罐 / 调料罐", "透明玻璃款"),
    ("Пепельница", "烟缸", "边界测试"),
    ("Аэратор для вина", "葡萄酒醒酒配件", "参考观察"),
    ("Графин", "分酒器 / 水壶", "透明玻璃款"),
    ("Чашка", "杯 / 咖啡茶杯", "部分相关"),
    ("Декантер", "醒酒器", "重点观察"),
]

CORE_GLASS_CATEGORIES = {
    "Стакан",
    "Бокал",
    "Чайник заварочный",
    "Кувшин",
    "Сахарница",
    "Графин",
    "Чашка",
    "Декантер",
}

SIDE_OBSERVATION_CATEGORIES = {
    "Крышка для кружки, чашки, стакана",
    "Пепельница",
    "Аэратор для вина",
    "Посуда для чайных церемоний",
}

CN_BY_CATEGORY = {key: (cn, role) for key, cn, role in CATEGORY_LABELS}
ORDER = {key: index for index, (key, _cn, _role) in enumerate(CATEGORY_LABELS)}


def as_number(value, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".")
    if text in {"—", "Нетданных", "Нет данных"}:
        return default
    try:
        return float(text)
    except ValueError:
        return default


def format_rmb_from_rub(value: object) -> str:
    amount = as_number(value) / RUB_PER_RMB
    if amount >= 100_000_000:
        return f"{amount / 100_000_000:.2f} 亿 RMB"
    if amount >= 10_000:
        return f"{amount / 10_000:.2f} 万 RMB"
    return f"{amount:.0f} RMB"


def format_int(value: object) -> str:
    return f"{int(round(as_number(value))):,}".replace(",", " ")


def format_percent(value: object) -> str:
    if value is None or str(value).strip() in {"", "—", "Нет данных"}:
        return "暂无"
    percent = as_number(value)
    if percent <= 1.5:
        percent *= 100
    return f"{percent:.1f}%"


def product_id(url: object) -> str:
    match = re.search(r"/product/(?:[^/]+-)?(\d+)", str(url))
    return match.group(1) if match else ""


def product_status(title: str, category: str) -> tuple[str, str]:
    text = title.lower()
    non_glass = [
        "фарфор",
        "керами",
        "глиня",
        "силикон",
        "металличес",
        "крышка",
        "пластик",
        "пластмас",
        "нержаве",
        "abs",
    ]
    if any(token in text for token in non_glass):
        return "非直接玻璃", "warn"
    direct_tokens = ["стекл", "бокал", "стакан", "графин", "декантер"]
    direct_categories = {"Стакан", "Бокал", "Чайник заварочный", "Графин", "Декантер"}
    if any(token in text for token in direct_tokens) or category in direct_categories:
        return "直接玻璃相关", "direct"
    if category in {"Пепельница", "Аэратор для вина", "Посуда для чайных церемоний", "Крышка для кружки, чашки, стакана"}:
        return "参考观察", "ref"
    return "待核材质", "ref"


def product_focus(title: str, category: str, status_class: str) -> tuple[str, str]:
    text = title.lower()
    non_glass_tokens = [
        "фарфор",
        "керами",
        "глиня",
        "силикон",
        "металличес",
        "крышка",
        "пластик",
        "пластмас",
        "нержаве",
        "abs",
    ]
    if any(token in text for token in non_glass_tokens):
        return "side", "旁类观察"
    if category in SIDE_OBSERVATION_CATEGORIES:
        return "side", "旁类观察"
    if status_class == "direct" and category in CORE_GLASS_CATEGORIES:
        return "core", "透明玻璃款"
    if "стекл" in text and category in CORE_GLASS_CATEGORIES:
        return "core", "透明玻璃款"
    if status_class == "direct":
        return "adjacent", "玻璃相关"
    return "side", "旁类观察"


def image_map() -> dict[str, str]:
    mapping: dict[str, str] = {}
    for path in ASSET_DIR.glob("ozon_*.*"):
        match = re.fullmatch(r"ozon_(\d+)\.(webp|jpg|jpeg|png)", path.name, re.IGNORECASE)
        if match:
            mapping[match.group(1)] = f"assets/ozon_product_images/{path.name}"
    return mapping


def is_product_sheet(ws) -> bool:
    link = ws.cell(7, 2).value
    title = ws.cell(7, 1).value
    return bool(title and link and "ozon.ru/product" in str(link))


def read_sections() -> list[dict]:
    workbook = load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    images = image_map()
    sections: list[dict] = []

    for ws in workbook.worksheets:
        if not is_product_sheet(ws):
            continue

        category = str(ws.cell(3, 2).value or ws.title)
        cn, role = CN_BY_CATEGORY.get(category, (category, "参考观察"))
        rows = []

        for row_number in range(7, min(ws.max_row, 16) + 1):
            title = ws.cell(row_number, 1).value
            link = ws.cell(row_number, 2).value
            if not title or not link:
                continue

            pid = product_id(link)
            avg_rub = as_number(ws.cell(row_number, 11).value)
            buyer_rmb = avg_rub / RUB_PER_RMB if avg_rub else 0
            status, status_class = product_status(str(title), category)
            focus, focus_label = product_focus(str(title), category, status_class)

            rows.append(
                {
                    "rank": len(rows) + 1,
                    "pid": pid,
                    "title": str(title),
                    "link": str(link),
                    "seller": str(ws.cell(row_number, 3).value or "暂无"),
                    "brand": str(ws.cell(row_number, 4).value or "暂无"),
                    "category": category,
                    "category_cn": cn,
                    "flag": str(ws.cell(row_number, 7).value or ""),
                    "sales_rub": as_number(ws.cell(row_number, 8).value),
                    "orders": as_number(ws.cell(row_number, 10).value),
                    "avg_rub": avg_rub,
                    "min_rub": as_number(ws.cell(row_number, 12).value),
                    "buyout": ws.cell(row_number, 13).value,
                    "image": images.get(pid, ""),
                    "status": status,
                    "status_class": status_class,
                    "focus": focus,
                    "focus_label": focus_label,
                    "normal_low": buyer_rmb / 0.7 if buyer_rmb else 0,
                    "normal_high": buyer_rmb / 0.5 if buyer_rmb else 0,
                }
            )

        if rows:
            sections.append(
                {
                    "key": category,
                    "cn": cn,
                    "role": role,
                    "sheet": ws.title,
                    "rows": rows,
                }
            )

    sections.sort(key=lambda item: ORDER.get(item["key"], 99))
    return sections


def sort_by_sales(items: list[dict]) -> list[dict]:
    return sorted(
        items,
        key=lambda item: (
            -item["sales_rub"],
            -item["orders"],
            item["rank"],
        ),
    )


def sort_by_orders(items: list[dict]) -> list[dict]:
    return sorted(
        items,
        key=lambda item: (
            -item["orders"],
            -item["sales_rub"],
            item["rank"],
        ),
    )


def with_display_rank(items: list[dict]) -> list[dict]:
    return [{**item, "display_rank": index} for index, item in enumerate(items, 1)]


def product_card(item: dict) -> str:
    if item["image"]:
        image_html = (
            f'<img src="{escape(item["image"])}" alt="{escape(item["title"])}" '
            f'data-lightbox-caption="{escape(item["title"])}">'
        )
    else:
        image_html = (
            '<div class="no-img">'
            "<b>待补 Ozon 主图</b>"
            f'<span>ID {escape(item["pid"] or "未识别")}</span>'
            "<small>不使用替图，避免错配</small>"
            "</div>"
        )

    buyer_rmb = item["avg_rub"] / RUB_PER_RMB if item["avg_rub"] else 0
    min_rmb = item["min_rub"] / RUB_PER_RMB if item["min_rub"] else 0
    normal_price = (
        f'{item["normal_low"]:.0f}-{item["normal_high"]:.0f} RMB'
        if item["normal_low"]
        else "待反推"
    )
    search_text = f'{item["title"]} {item["seller"]} {item["brand"]}'.lower()

    rank = item.get("display_rank", item["rank"])

    return f"""
      <article class="product-card" data-status="{item['status_class']}" data-focus="{item['focus']}" data-has-img="{'1' if item['image'] else '0'}" data-sales="{int(round(item['sales_rub']))}" data-orders="{int(round(item['orders']))}" data-search="{escape(search_text)}">
        <a class="thumb" href="{escape(item['link'])}" target="_blank" rel="noopener noreferrer">{image_html}</a>
        <div class="body">
          <div class="topline">
            <span class="rank">#{rank:02d}</span>
            <span class="badge {item['focus']}">{escape(item['focus_label'])}</span>
          </div>
          <h3>{escape(item['title'])}</h3>
          <div class="meta">
            <span>类目：<b>{escape(item['category_cn'])}</b></span>
            <span>卖家：<b>{escape(item['seller'])}</b></span>
            <span>品牌：<b>{escape(item['brand'])}</b></span>
          </div>
          <div class="metrics">
            <div><b>{format_rmb_from_rub(item['sales_rub'])}</b><span>28天销售额</span></div>
            <div><b>{format_int(item['orders'])}</b><span>28天订单</span></div>
            <div><b>{buyer_rmb:.0f} RMB</b><span>买家均价 / 原始 {format_int(item['avg_rub'])} ₽</span></div>
            <div><b>{normal_price}</b><span>反推平台标价</span></div>
          </div>
          <div class="note">最低买家价：{min_rmb:.0f} RMB（原始 {format_int(item['min_rub'])} ₽） · 签收率：{format_percent(item['buyout'])}</div>
          <a class="open" href="{escape(item['link'])}" target="_blank" rel="noopener noreferrer">打开 Ozon 商品 ↗</a>
        </div>
      </article>
    """


def render(sections: list[dict]) -> str:
    all_items = [item for section in sections for item in section["rows"]]
    core_items = with_display_rank(sort_by_orders([item for item in all_items if item["focus"] == "core"]))
    verified_images = sum(1 for item in all_items if item["image"])
    core_images = sum(1 for item in core_items if item["image"])
    side_count = sum(1 for item in all_items if item["focus"] != "core")

    nav_html = "".join(
        (
            f'<button class="nav-pill" data-category="{escape(section["key"])}">'
            f'<b>{escape(section["cn"])}</b><span>{len(section["rows"])} 款</span></button>'
        )
        for section in sections
    )
    nav_html = (
        f'<button class="nav-pill active" data-category="core"><b>透明玻璃款</b>'
        f'<span>{len(core_items)} 款</span></button>'
        + nav_html
    )
    core_html = f"""
        <section class="cat-section core-wall" id="core-glass" data-section-kind="core" data-category="core">
          <div class="cat-head">
            <div>
              <div class="eyebrow">MAIN GLASS ITEMS · 只看德力最相关的透明玻璃款</div>
              <h2>透明玻璃款</h2>
              <p>这里不再平均展示所有类目，而是把水杯、高脚杯、玻璃茶壶、水壶、分酒器、醒酒器、糖罐等与德力产品线直接相关的透明玻璃款集中放前面。杯盖、烟缸、硅胶、金属、陶瓷茶具只放到后面的旁类观察。</p>
            </div>
            <div class="cat-stats">
              <div><b>{format_rmb_from_rub(sum(item['sales_rub'] for item in core_items))}</b><span>透明玻璃样本销售额</span></div>
              <div><b>{format_int(sum(item['orders'] for item in core_items))}</b><span>透明玻璃样本订单</span></div>
              <div><b>{core_images}/{len(core_items)}</b><span>已核准 Ozon 主图</span></div>
            </div>
          </div>
          <div class="focus-note"><b>展示口径：</b>优先用于给工厂讲“德力该切哪里”。价格仍按买家成交价展示，并反推平台标价；主图必须与 Ozon 商品 ID 对上。</div>
          <div class="cards core-cards">{''.join(product_card(item) for item in core_items)}</div>
        </section>
        """
    section_html = ""
    for section in sections:
        rows = with_display_rank(sort_by_orders(section["rows"]))
        section_html += f"""
        <section class="cat-section" id="{escape(section['key'])}" data-section-kind="category" data-category="{escape(section['key'])}">
          <div class="cat-head">
            <div>
              <div class="eyebrow">{escape(section['role'])} · {escape(section['key'])}</div>
              <h2>{escape(section['cn'])}</h2>
              <p>本类目默认按 28 天销量从高到低展示 Excel 热卖商品榜前 10，也可以切换为销售额排序。图片只放已核准 Ozon 商品 ID 的本地素材；缺图商品保留链接和数据，不用替图。</p>
            </div>
            <div class="cat-stats">
              <div><b>{format_rmb_from_rub(sum(item['sales_rub'] for item in rows))}</b><span>前10合计销售额</span></div>
              <div><b>{format_int(sum(item['orders'] for item in rows))}</b><span>前10合计订单</span></div>
              <div><b>{sum(1 for item in rows if item['image'])}/10</b><span>已核准主图</span></div>
            </div>
          </div>
          <div class="cards">{''.join(product_card(item) for item in rows)}</div>
        </section>
        """

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Ozon 竞品库 · 德力玻璃俄罗斯电商</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,300;9..144,500&family=Inter:wght@400;500;600;700&family=Noto+Sans+SC:wght@400;500;700&display=swap" rel="stylesheet">
<style>
:root{{--bg:#11100e;--panel:#191612;--paper:#f3eadb;--muted:#b9aa94;--soft:#7e705f;--line:#3a3127;--copper:#d58a45;--green:#7fa76a;--red:#d06b55;--blue:#77a6c8}}
*{{box-sizing:border-box}}html{{scroll-behavior:smooth}}body{{margin:0;background:var(--bg);color:var(--paper);font-family:Inter,"Noto Sans SC",sans-serif;line-height:1.6}}a{{color:inherit}}.wrap{{max-width:1480px;margin:auto;padding:0 clamp(18px,4vw,56px)}}
header{{padding:34px 0 28px;border-bottom:1px solid var(--line);position:sticky;top:0;background:rgba(17,16,14,.92);backdrop-filter:blur(14px);z-index:20}}.top{{display:flex;justify-content:space-between;gap:20px;align-items:center}}.brand{{font-family:Fraunces,serif;font-size:22px}}.back{{text-decoration:none;border:1px solid var(--line);padding:9px 13px;color:var(--muted);font-size:13px}}.back:hover{{border-color:var(--copper);color:var(--paper)}}
.hero{{padding:56px 0 36px}}h1{{font-family:Fraunces,"Noto Sans SC",serif;font-weight:300;font-size:clamp(42px,7vw,96px);line-height:.98;margin:0 0 22px}}.lead{{max-width:900px;color:var(--muted);font-size:17px}}.summary{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-top:30px}}.summary div,.cat-stats div{{border:1px solid var(--line);background:var(--panel);padding:18px}}.summary b{{display:block;font-family:Fraunces,serif;font-size:32px;font-weight:300}}.summary span,.cat-stats span{{color:var(--muted);font-size:13px}}
.toolbar{{position:sticky;top:97px;z-index:15;background:rgba(17,16,14,.95);border-top:1px solid var(--line);border-bottom:1px solid var(--line);padding:14px 0}}.tools{{display:grid;grid-template-columns:1fr auto;gap:18px;align-items:center}}input{{width:100%;background:var(--panel);border:1px solid var(--line);color:var(--paper);padding:12px 14px;font-size:14px}}.filters{{display:flex;gap:8px;flex-wrap:wrap}}button{{background:var(--panel);border:1px solid var(--line);color:var(--muted);padding:10px 12px;cursor:pointer}}button.active,button:hover{{border-color:var(--copper);color:var(--paper)}}button.active{{background:linear-gradient(135deg,rgba(213,138,69,.26),rgba(25,22,18,.96));box-shadow:0 10px 28px rgba(213,138,69,.18)}}.nav{{display:flex;gap:8px;overflow:auto;padding:14px 0 0;scrollbar-width:none;-ms-overflow-style:none}}.nav::-webkit-scrollbar{{display:none}}.nav-pill{{min-width:150px;text-align:left;text-decoration:none;border:1px solid var(--line);background:var(--panel);padding:10px 12px}}.nav-pill.active{{border-color:var(--copper);color:var(--paper);background:rgba(213,138,69,.08)}}.nav b{{display:block;font-size:13px}}.nav span{{color:var(--soft);font-size:12px}}
.sort-note{{margin-top:10px;color:var(--muted);font-size:12px;display:flex;gap:10px;align-items:center;flex-wrap:wrap}}.sort-note b{{color:var(--paper);font-weight:600}}.sort-state{{display:inline-flex;align-items:center;min-height:24px;padding:3px 8px;border:1px solid rgba(213,138,69,.42);background:rgba(213,138,69,.08);color:var(--paper);font-weight:600}}
.notice{{margin:28px 0 10px;border:1px solid rgba(213,138,69,.35);background:rgba(213,138,69,.08);padding:18px;color:var(--muted)}}.notice b{{color:var(--paper)}}.notice ul{{margin:8px 0 0 20px}}
.cat-section{{padding:64px 0;border-bottom:1px solid var(--line)}}.cat-head{{display:grid;grid-template-columns:1fr auto;gap:30px;align-items:end;margin-bottom:22px}}.eyebrow{{color:var(--copper);font-size:12px;letter-spacing:.12em;text-transform:uppercase}}h2{{font-family:Fraunces,"Noto Sans SC",serif;font-size:44px;font-weight:300;margin:8px 0}}.cat-head p{{margin:0;color:var(--muted)}}.cat-stats{{display:grid;grid-template-columns:repeat(3,150px);gap:10px}}.cat-stats b{{display:block;font-size:18px}}
.cards{{display:grid;grid-template-columns:repeat(5,1fr);gap:14px}}.product-card{{background:var(--panel);border:1px solid var(--line);display:flex;flex-direction:column;min-height:620px}}.thumb{{display:grid;place-items:center;aspect-ratio:1/1;background:#f8f4ec;border-bottom:1px solid var(--line);text-decoration:none;position:relative;overflow:hidden}}.thumb img{{width:100%;height:100%;object-fit:contain;padding:14px;cursor:zoom-in}}.no-img{{color:#6d6253;text-align:center;padding:20px}}.no-img b{{display:block;color:#332b22}}.no-img span,.no-img small{{display:block;margin-top:6px}}.body{{padding:16px;display:flex;flex-direction:column;gap:12px;flex:1}}.topline{{display:flex;justify-content:space-between;gap:8px}}.rank{{font-family:Fraunces,serif;color:var(--copper);font-size:22px}}.badge{{font-size:11px;border:1px solid currentColor;padding:4px 7px;height:max-content}}.badge.core{{color:var(--green)}}.badge.side{{color:var(--blue)}}.badge.direct{{color:var(--green)}}.badge.warn{{color:var(--red)}}.badge.ref{{color:var(--blue)}}h3{{font-size:15px;line-height:1.45;margin:0;display:-webkit-box;-webkit-line-clamp:4;-webkit-box-orient:vertical;overflow:hidden}}.meta{{color:var(--muted);font-size:12px;display:grid;gap:3px}}.meta b{{color:var(--paper);font-weight:500}}.metrics{{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:auto}}.metrics div{{border:1px solid var(--line);background:rgba(255,255,255,.025);padding:9px}}.metrics b{{display:block;font-size:14px}}.metrics span,.note{{color:var(--muted);font-size:11px}}.open{{display:block;text-align:center;text-decoration:none;border:1px solid var(--line);padding:10px;color:var(--copper);font-size:13px}}.open:hover{{background:rgba(213,138,69,.08)}}
.lightbox{{position:fixed;inset:0;background:rgba(0,0,0,.82);display:none;place-items:center;z-index:100;padding:28px}}.lightbox.open{{display:grid}}.lightbox img{{max-width:92vw;max-height:82vh;background:#fff}}.cap{{color:var(--paper);margin-top:12px;max-width:900px;text-align:center}}.close{{position:absolute;right:20px;top:20px;border-color:#777;color:#fff}}.hidden{{display:none!important}}footer{{padding:38px 0;color:var(--soft);font-size:12px}}
@media(max-width:1200px){{.cards{{grid-template-columns:repeat(3,1fr)}}.summary{{grid-template-columns:repeat(2,1fr)}}}}@media(max-width:760px){{.cards,.summary,.cat-head,.tools{{grid-template-columns:1fr}}.cat-stats{{grid-template-columns:1fr 1fr 1fr}}header{{position:static}}.toolbar{{top:0}}}}
</style>
</head>
<body>
<header><div class="wrap top"><div class="brand">Ozon 竞品库 · 德力俄罗斯项目</div><a class="back" href="index.html">返回合作提案</a></div></header>
<main>
  <div class="wrap hero">
    <h1>先看透明玻璃，<br>再看旁类趋势。</h1>
    <p class="lead">这个页面专门承接 Ozon 竞品。默认只展示和德力最相关的透明玻璃商品，并按 28 天销量从高到低排，方便先看谁真正卖得动；热卖商品价格按买家价 / 活动价口径展示，反推平台标价按“买家价 ÷ 0.7 到 ÷ 0.5”估算。没有核准到商品 ID 的图片不使用替图，只保留 Ozon 链接和数据。</p>
    <div class="summary"><div><b>{len(core_items)}</b><span>个透明玻璃款样本</span></div><div><b>{len(all_items)}</b><span>款全量热卖样本</span></div><div><b>{verified_images}</b><span>张已核准主图</span></div><div><b>{side_count}</b><span>款旁类观察样本</span></div></div>
    <div class="notice"><b>展示逻辑已经收窄</b><ul><li><b>透明玻璃款</b>：水杯、高脚杯、玻璃茶壶、水壶、分酒器、醒酒器、糖罐、透明玻璃杯。</li><li><b>旁类观察</b>：杯盖、烟缸、硅胶、金属、陶瓷茶具等只用来观察平台内容与价格打法，不作为德力透明玻璃款的直接对标样本。</li><li><b>数据缺口</b>：当前 Excel 未提供真正的马克杯 28 天热卖榜；现有“Кружки28天热卖”实际是杯盖配件。</li></ul></div>
  </div>
  <div class="toolbar"><div class="wrap"><div class="tools"><input id="q" placeholder="搜索商品、品牌、卖家，例如 Pasabahce / ThermoGlass / 玻璃茶壶"><div class="filters"><button class="active" data-filter="core" data-sort="orders">透明玻璃款</button><button data-filter="top-orders" data-sort="orders">按销量排序</button><button data-filter="top-sales" data-sort="sales">按销售额排序</button><button data-filter="all">全部类目</button><button data-filter="side">旁类观察</button><button data-filter="direct">直接玻璃</button><button data-filter="img">已有主图</button></div></div><div class="sort-note"><b>排序口径：</b><span id="sort-state" class="sort-state">当前：按 28 天销量从高到低</span><span>切换类目、筛选或搜索后仍保持当前排序。</span></div><nav class="nav">{nav_html}</nav></div></div>
  <div class="wrap">{core_html}{section_html}</div>
</main>
<footer><div class="wrap">数据来源：重庆得力玻璃类目趋势.xlsx · Ozon 28 天热卖商品榜 · 人民币折算按 1 RMB = 12.8 RUB。图片口径：只展示本地已按 Ozon 商品 ID 命名并与链接匹配的素材。</div></footer>
<div class="lightbox" id="lightbox"><button class="close">关闭</button><div><img alt=""><div class="cap"></div></div></div>
<script>
const q=document.getElementById('q');
const buttons=[...document.querySelectorAll('button[data-filter]')];
const cards=[...document.querySelectorAll('.product-card')];
const sections=[...document.querySelectorAll('.cat-section')];
const navPills=[...document.querySelectorAll('.nav-pill')];
let filter='core';
let category='core';
let sortMode='orders';
function sortCards(){{
  sections.forEach(section=>{{
    section.querySelectorAll('.cards').forEach(holder=>{{
      [...holder.querySelectorAll('.product-card')]
        .sort((a,b)=>{{
          const primary=sortMode==='sales'?'sales':'orders';
          const secondary=sortMode==='sales'?'orders':'sales';
          return (Number(b.dataset[primary])||0)-(Number(a.dataset[primary])||0)||((Number(b.dataset[secondary])||0)-(Number(a.dataset[secondary])||0));
        }})
        .forEach(card=>holder.appendChild(card));
      [...holder.querySelectorAll('.product-card')].forEach((card,index)=>{{
        const rank=card.querySelector('.rank');
        if(rank) rank.textContent=`#${{String(index+1).padStart(2,'0')}}`;
      }});
    }});
  }});
}}
function apply(){{
  sortCards();
  document.getElementById('sort-state').textContent=sortMode==='sales'?'当前：按 28 天销售额从高到低':'当前：按 28 天销量从高到低';
  const term=q.value.trim().toLowerCase();
  cards.forEach(card=>{{
    let ok=!term||card.dataset.search.includes(term);
    if(category==='core') ok=ok&&card.dataset.focus==='core';
    else if(category!=='all') ok=ok&&card.closest('.cat-section')?.dataset.category===category;
    if(filter==='core'||filter==='top-orders'||filter==='top-sales') ok=ok&&card.dataset.focus==='core';
    if(filter==='side') ok=ok&&card.dataset.focus!=='core';
    if(filter==='direct') ok=ok&&card.dataset.status==='direct';
    if(filter==='img') ok=ok&&card.dataset.hasImg==='1';
    card.classList.toggle('hidden',!ok);
  }});
  sections.forEach(section=>{{
    if(category==='core'&&section.dataset.sectionKind!=='core'){{section.classList.add('hidden');return;}}
    if(category==='all'&&section.dataset.sectionKind==='core'){{section.classList.add('hidden');return;}}
    if(category!=='core'&&category!=='all'&&section.dataset.category!==category){{section.classList.add('hidden');return;}}
    section.classList.toggle('hidden',!section.querySelector('.product-card:not(.hidden)'));
  }});
}}
q.addEventListener('input',apply);
buttons.forEach(button=>button.addEventListener('click',()=>{{buttons.forEach(item=>item.classList.remove('active'));button.classList.add('active');filter=button.dataset.filter;if(button.dataset.sort) sortMode=button.dataset.sort;category=(filter==='core'||filter==='top-orders'||filter==='top-sales')?'core':'all';navPills.forEach(item=>item.classList.toggle('active',item.dataset.category===category));apply();}}));
navPills.forEach(button=>button.addEventListener('click',()=>{{navPills.forEach(item=>item.classList.remove('active'));button.classList.add('active');category=button.dataset.category;filter=category==='core'?'core':'all';buttons.forEach(item=>item.classList.toggle('active',item.dataset.filter===filter));apply();const target=document.querySelector(`.cat-section[data-category="${{CSS.escape(category)}}"]`);target?.scrollIntoView({{behavior:'smooth',block:'start'}});}}));
const lb=document.getElementById('lightbox'),lbImg=lb.querySelector('img'),cap=lb.querySelector('.cap');
document.addEventListener('click',event=>{{const img=event.target.closest('.thumb img');if(!img)return;event.preventDefault();lbImg.src=img.src;cap.textContent=img.dataset.lightboxCaption||img.alt;lb.classList.add('open');}});
lb.addEventListener('click',event=>{{if(event.target===lb||event.target.classList.contains('close'))lb.classList.remove('open');}});
window.addEventListener('keydown',event=>{{if(event.key==='Escape')lb.classList.remove('open');}});
apply();
</script>
</body>
</html>
"""


def main() -> None:
    sections = read_sections()
    html = render(sections)
    OUT_HTML.write_text(html, encoding="utf-8")
    all_items = [item for section in sections for item in section["rows"]]
    print(
        json.dumps(
            {
                "out": str(OUT_HTML),
                "sections": len(sections),
                "items": len(all_items),
                "verified_images": sum(1 for item in all_items if item["image"]),
                "core_glass": sum(1 for item in all_items if item["focus"] == "core"),
                "side_observation": sum(1 for item in all_items if item["focus"] != "core"),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
